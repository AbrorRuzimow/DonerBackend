from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse, reverse_lazy
import openpyxl
from openpyxl.workbook import Workbook
from App.excel import *
from App.models import *
from App.views import BaseListView, BaseCreateView, BaseUpdateView, BaseView


class HomePageList(BaseListView):
    template_name = 'administrator/home_page/list.html'

    def get_queryset(self):
        models = HomePage.objects.all()
        return models

    def get_context_data(self, **kwargs):
        context = super(HomePageList, self).get_context_data(**kwargs)
        context['models_all'] = self.get_queryset().count()
        return context


class HomePageCreate(BaseCreateView):
    template_name = 'administrator/home_page/create.html'
    model = HomePage
    fields = '__all__'
    success_url = reverse_lazy('Administrator:homepage_list')
    success_message = '123'


class HomePageUpdate(BaseUpdateView):
    template_name = 'administrator/home_page/update.html'
    model = HomePage
    fields = '__all__'
    success_url = reverse_lazy('Administrator:homepage_list')


class HomePageDelete(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        try:
            HomePage.objects.get(pk=kwargs['pk']).delete()
            messages.success(request, "Maglumat pozuldy")
        except ObjectDoesNotExist:
            messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:homepage_list'))


class HomePageMultiDelete(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        for i in request.POST.getlist('item_id'):
            try:
                HomePage.objects.get(pk=i).delete()
                messages.success(request, "Maglumat pozuldy")
            except ObjectDoesNotExist:
                messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:homepage_list'))


class HomePageExcelDownload(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="HomePage.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = 'HomePage'
        sheet = wb.worksheets[0]
        # Ady
        cell = sheet.cell(row=1, column=1, value='Ady')
        cell.font = font_14_bold
        cell.border = border
        cell.alignment = center
        cell.fill = bg_color_light_green
        sheet.row_dimensions[1].height = 20
        sheet.column_dimensions[Alphabet(1)].width = 16
        wb.save(response)
        return response


class HomePageExcelUpload(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        try:
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            shell = wb.active
            c = True
            i = 1
            while c:
                i += 1
                if shell.cell(row=i, column=1).value is None:
                    c = False
                    break
                else:
                    model = HomePage()
                    model.name = shell.cell(row=i, column=1).value
                    model.save()
                messages.success(request, 'Maglumat girizildi')
            return HttpResponseRedirect(reverse('Administrator:warehouse_name_list'))
        except Exception as e:
            messages.error(request, f'Maglumat girizilmedi: {e}')
            return HttpResponseRedirect(reverse('Administrator:warehouse_name_list'))
