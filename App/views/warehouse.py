from datetime import datetime

from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse, reverse_lazy
import openpyxl
from openpyxl.workbook import Workbook
from App.excel import *
from App.models import *
from App.views import BaseListView, BaseCreateView, BaseUpdateView, BaseView


class WarehouseList(BaseListView):
    template_name = 'administrator/warehouse/list.html'

    def get_queryset(self):
        models = Warehouse.objects.all().order_by('-date', 'warehouse_name_fk__name')
        if self.request.GET.get('warehouse'):
            models = models.filter(warehouse_name_fk__in=self.request.GET.get('warehouse').split(','))
        if self.request.GET.get('date') and self.request.GET.get('date') != 'None':
            date = self.request.GET.get('date').split('to')
            date1 = date[0].split(' ')[0]
            date2 = date[1].split(' ')[1]
            date_start = datetime.strptime(date1, '%Y-%m-%d')
            date_end = datetime.strptime(date2, '%Y-%m-%d')
            models = models.filter(date__gte=date_start, date__lte=date_end)
        if self.request.GET.get('status'):
            if self.request.GET.get('status') == '0':
                pass
            else:
                models = models.filter(status=self.request.GET.get('status'))
        return models

    def get_context_data(self, **kwargs):
        context = super(WarehouseList, self).get_context_data(**kwargs)
        context['warehouse_get'] = self.request.GET.get('warehouse', '').strip(',')
        context['date_get'] = self.request.GET.get('date', '')
        context['status_get'] = self.request.GET.get('status', '')
        context['models_all'] = self.get_queryset().count()
        context['status_list'] = Warehouse.status_list
        warehouse = WarehouseName.objects.all().order_by('name')
        context['warehouse_name'] = warehouse
        warehouse_model_list = []
        for i in warehouse:
            w_list = {}
            models = self.get_queryset().filter(warehouse_name_fk=i)
            if models.count() > 0:
                w_list['name'] = i.name
                amount = 0
                for model in models:
                    amount = amount + model.amount
                w_list['amount'] = amount
                amount_use = 0
                for model in models:
                    amount_use = amount_use + model.amount_use
                w_list['amount_use'] = amount_use
                price = 0
                for model in models:
                    price = price + model.price
                w_list['price'] = price
                warehouse_model_list.append(w_list)
        context['warehouse_model_list'] = warehouse_model_list
        return context


class WarehouseCreate(BaseCreateView):
    template_name = 'administrator/warehouse/create.html'
    model = Warehouse
    fields = ('warehouse_name_fk', 'amount', 'price',)
    success_url = reverse_lazy('Administrator:warehouse_list')

    def get_context_data(self, **kwargs):
        context = super(WarehouseCreate, self).get_context_data(**kwargs)
        context['warehouse_name_models'] = WarehouseName.objects.all()
        return context


class WarehouseUpdate(BaseUpdateView):
    template_name = 'administrator/warehouse/update.html'
    model = Warehouse
    fields = ('warehouse_name_fk', 'amount', 'price',)
    success_url = reverse_lazy('Administrator:warehouse_list')

    def get_context_data(self, **kwargs):
        context = super(WarehouseUpdate, self).get_context_data(**kwargs)
        context['warehouse_name_models'] = WarehouseName.objects.all()
        return context


class WarehouseDelete(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        try:
            Warehouse.objects.get(pk=kwargs['pk']).delete()
            messages.success(request, "Maglumat pozuldy")
        except ObjectDoesNotExist:
            messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:warehouse_list'))


class WarehouseMultiDelete(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        for i in request.POST.getlist('item_id'):
            try:
                Warehouse.objects.get(pk=i).delete()
                messages.success(request, "Maglumat pozuldy")
            except ObjectDoesNotExist:
                messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:warehouse_list'))


class WarehouseExcelDownload(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Warehouse.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = 'Warehouse'
        sheet = wb.worksheets[0]
        # Ady
        cell = sheet.cell(row=1, column=1, value='Ady')
        cell.font = font_14_bold
        cell.border = border
        cell.alignment = center
        cell.fill = bg_color_light_green
        sheet.row_dimensions[1].height = 20
        sheet.column_dimensions[Alphabet(1)].width = 16
        wb.save(response)
        return response


class WarehouseExcelUpload(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        try:
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            shell = wb.active
            c = True
            i = 1
            while c:
                i += 1
                if shell.cell(row=i, column=1).value is None:
                    c = False
                    break
                else:
                    model = Warehouse()
                    model.name = shell.cell(row=i, column=1).value
                    model.save()
                messages.success(request, 'Maglumat girizildi')
            return HttpResponseRedirect(reverse('Administrator:warehouse_list'))
        except Exception as e:
            messages.error(request, f'Maglumat girizilmedi: {e}')
            return HttpResponseRedirect(reverse('Administrator:warehouse_list'))
