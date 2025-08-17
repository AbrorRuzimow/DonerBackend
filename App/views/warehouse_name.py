from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponseRedirect, HttpResponse
from django.urls import reverse, reverse_lazy
import openpyxl
from openpyxl.workbook import Workbook
from App.excel import *
from App.models import *
from App.views import BaseListView, BaseCreateView, BaseUpdateView, BaseView


class WarehouseNameList(BaseListView):
    template_name = 'administrator/warehouse_name/list.html'

    def get_queryset(self):
        models = WarehouseName.objects.all().order_by('name')
        if self.request.GET.get('search'):
            models = models.filter(name__icontains=self.request.GET.get('search'))
        if self.request.GET.get('order_by'):
            models = models.order_by(self.request.GET.get('order_by'))
        return models

    def get_context_data(self, **kwargs):
        context = super(WarehouseNameList, self).get_context_data(**kwargs)
        context['order_by'] = self.request.GET.get('order_by', 'name')
        context['models_all'] = self.get_queryset().count()
        return context


class WarehouseNameCreate(BaseCreateView):
    template_name = 'administrator/warehouse_name/create.html'
    model = WarehouseName
    fields = '__all__'
    success_url = reverse_lazy('Administrator:warehouse_name_list')
    success_message = '123'


class WarehouseNameUpdate(BaseUpdateView):
    template_name = 'administrator/warehouse_name/update.html'
    model = WarehouseName
    fields = '__all__'
    success_url = reverse_lazy('Administrator:warehouse_name_list')



class WarehouseNameDelete(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        try:
            WarehouseName.objects.get(pk=kwargs['pk']).delete()
            messages.success(request, "Maglumat pozuldy")
        except ObjectDoesNotExist:
            messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:warehouse_name_list'))


class WarehouseNameMultiDelete(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        for i in request.POST.getlist('item_id'):
            try:
                WarehouseName.objects.get(pk=i).delete()
                messages.success(request, "Maglumat pozuldy")
            except ObjectDoesNotExist:
                messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:warehouse_name_list'))


class WarehouseNameExcelDownload(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="WarehouseName.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = 'WarehouseName'
        sheet = wb.worksheets[0]
        # Ady
        cell = sheet.cell(row=1, column=1, value='Ady')
        cell.font = font_14_bold
        cell.border = border
        cell.alignment = center
        cell.fill = bg_color_light_green
        sheet.row_dimensions[1].height = 20
        sheet.column_dimensions[Alphabet(1)].width = 16
        wb.save(response)
        return response


class WarehouseNameExcelUpload(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        try:
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            shell = wb.active
            c = True
            i = 1
            while c:
                i += 1
                if shell.cell(row=i, column=1).value is None:
                    c = False
                    break
                else:
                    model = WarehouseName()
                    model.name = shell.cell(row=i, column=1).value
                    model.save()
                messages.success(request, 'Maglumat girizildi')
            return HttpResponseRedirect(reverse('Administrator:warehouse_name_list'))
        except Exception as e:
            messages.error(request, f'Maglumat girizilmedi: {e}')
            return HttpResponseRedirect(reverse('Administrator:warehouse_name_list'))
