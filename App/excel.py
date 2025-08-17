from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

font_14_bold = Font(
    name='Times New Roman',
    size=14,
    bold=True
)
font_14 = Font(
    name='Times New Roman',
    size=14,
)
border = Border(
    left=Side(border_style='thin', color='00000000'),
    right=Side(border_style='thin', color='00000000'),
    top=Side(border_style='thin', color='00000000'),
    bottom=Side(border_style='thin', color='00000000'),
)
center = Alignment(
    horizontal='center',
    vertical='center',
    wrap_text=True,
)
bg_color_light_green = PatternFill(
    fgColor='30e998',
    fill_type='solid'
)

bg_color_light_yellow = PatternFill(
    fgColor='0fd4e1',
    fill_type='solid'
)


def Alphabet(a):
    if 1 <= a <= 26:
        letter = f"{chr(a + 64)}"
        return letter
    elif 27 <= a <= 52:
        letter = f"A{chr(a + 38)}"
        return letter
    elif 53 <= a <= 78:
        letter = f"B{chr(a + 38)}"
        return letter
    return None
