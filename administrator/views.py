from datetime import datetime

import openpyxl
from django.contrib import messages
from django.contrib.auth import logout, authenticate, login
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import ExpressionWrapper, F, DecimalField, OuterRef, Subquery, Value, Case, When, IntegerField
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.urls import reverse, reverse_lazy
from django.views import View
from django.views.generic import TemplateView, ListView, CreateView, UpdateView
from openpyxl import Workbook

from Api.models import *
from administrator.excel import *


class LoginView(TemplateView):
    template_name = 'login.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            if request.POST.get('next', None):
                return HttpResponseRedirect(request.POST['next'])
            if request.user.user_type == '1':
                return HttpResponseRedirect(reverse('Administrator:dashboard'))
            elif request.user.user_type == '2':
                return HttpResponseRedirect(reverse('Manager:dashboard'))
            elif request.user.user_type == '3':
                return HttpResponseRedirect(reverse('App:home'))
            return HttpResponseRedirect(reverse('logout'))
        else:
            return super().dispatch(request, *args, **kwargs)

    @staticmethod
    def post(request, *args, **kwargs):
        user = authenticate(request, username=request.POST['username'], password=request.POST['password'])
        if user is not None:
            login(request, user)
            print(request.user.user_type)
            if request.POST.get('next', None):
                return HttpResponseRedirect(request.POST['next'])
            if request.user.user_type == '1':
                return HttpResponseRedirect(reverse('Administrator:dashboard'))
            elif request.user.user_type == '2':
                return HttpResponseRedirect(reverse('Manager:dashboard'))
            return HttpResponseRedirect(reverse('Authenticate:logout'))

        else:
            messages.error(request, 'Username or Password is incorrect')
            return HttpResponseRedirect(reverse('Authenticate:login'))


def logout_view(request):
    logout(request)
    return HttpResponseRedirect(reverse('login'))


class BaseTemplateView(TemplateView):
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.is_superuser and request.user.user_type == '1':
            return super().dispatch(request, *args, **kwargs)
        else:
            return HttpResponseRedirect(reverse('Authenticate:logout'))


class BaseListView(ListView):
    context_object_name = 'models'
    paginate_by = 10
    model = None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page'] = self.request.GET.get('page', '1')
        context['search'] = self.request.GET.get('search', '')
        context['order_by'] = self.request.GET.get('order_by', '')
        context['limit'] = self.request.GET.get('limit', 10)
        return context

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.is_superuser and request.user.user_type == '1':
            return super().dispatch(request, *args, **kwargs)
        else:
            return HttpResponseRedirect(reverse('Authenticate:logout'))

    def get_paginate_by(self, queryset):
        limit_page = self.request.GET.get('limit', 10)
        return limit_page


class BaseCreateView(CreateView):
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.is_superuser and request.user.user_type == '1':
            return super().dispatch(request, *args, **kwargs)
        else:
            return HttpResponseRedirect(reverse('Authenticate:logout'))

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, "Maglumat girizildi")
        return response

    def form_invalid(self, form):
        response = super().form_invalid(form)
        messages.error(self.request, f"Ýalňyşlyk ýüze çykdy: {form.errors}")
        return response


class BaseUpdateView(UpdateView):
    context_object_name = 'model'

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, "Maglumat üýtgedildi")
        return response

    def form_invalid(self, form):
        response = super().form_invalid(form)
        messages.success(self.request, "Ýalňyşlyk ýüze çykdy")
        return response

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.is_superuser and request.user.user_type == '1':
            return super().dispatch(request, *args, **kwargs)
        else:
            return HttpResponseRedirect(reverse('Authenticate:logout'))


class BaseView(View):
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and request.user.is_superuser and request.user.user_type == '1':
            return super().dispatch(request, *args, **kwargs)
        else:
            return HttpResponseRedirect(reverse('Authenticate:logout'))


class Dashboard(BaseTemplateView):
    template_name = 'administrator/dashboard/index.html'


class WarehouseNameList(BaseListView):
    template_name = 'administrator/warehouse_name/list.html'

    def get_queryset(self):
        models = WarehouseName.objects.all().order_by('name')
        if self.request.GET.get('search'):
            models = models.filter(name__icontains=self.request.GET.get('search'))
        if self.request.GET.get('order_by'):
            models = models.order_by(self.request.GET.get('order_by'))
        return models

    def get_context_data(self, **kwargs):
        context = super(WarehouseNameList, self).get_context_data(**kwargs)
        context['order_by'] = self.request.GET.get('order_by', 'name')
        context['models_all'] = self.get_queryset().count()
        return context


class WarehouseNameCreate(BaseCreateView):
    template_name = 'administrator/warehouse_name/create.html'
    model = WarehouseName
    fields = '__all__'
    success_url = reverse_lazy('Administrator:warehouse_name_list')
    success_message = '123'


class WarehouseNameUpdate(BaseUpdateView):
    template_name = 'administrator/warehouse_name/update.html'
    model = WarehouseName
    fields = '__all__'
    success_url = reverse_lazy('Administrator:warehouse_name_list')


class WarehouseNameDelete(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        try:
            WarehouseName.objects.get(pk=kwargs['pk']).delete()
            messages.success(request, "Maglumat pozuldy")
        except ObjectDoesNotExist:
            messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:warehouse_name_list'))


class WarehouseNameMultiDelete(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        for i in request.POST.getlist('item_id'):
            try:
                WarehouseName.objects.get(pk=i).delete()
                messages.success(request, "Maglumat pozuldy")
            except ObjectDoesNotExist:
                messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:warehouse_name_list'))


class WarehouseNameExcelDownload(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="WarehouseName.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = 'WarehouseName'
        sheet = wb.worksheets[0]
        # Ady
        cell = sheet.cell(row=1, column=1, value='Ady')
        cell.font = font_14_bold
        cell.border = border
        cell.alignment = center
        cell.fill = bg_color_light_green
        sheet.row_dimensions[1].height = 20
        sheet.column_dimensions[Alphabet(1)].width = 16
        wb.save(response)
        return response


class WarehouseNameExcelUpload(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        try:
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            shell = wb.active
            c = True
            i = 1
            while c:
                i += 1
                if shell.cell(row=i, column=1).value is None:
                    c = False
                    break
                else:
                    model = WarehouseName()
                    model.name = shell.cell(row=i, column=1).value
                    model.save()
                messages.success(request, 'Maglumat girizildi')
            return HttpResponseRedirect(reverse('Administrator:warehouse_name_list'))
        except Exception as e:
            messages.error(request, f'Maglumat girizilmedi: {e}')
            return HttpResponseRedirect(reverse('Administrator:warehouse_name_list'))


class WarehouseList(BaseListView):
    template_name = 'administrator/warehouse/list.html'

    def get_queryset(self):
        models = Warehouse.objects.all().order_by('-date', 'warehouse_name_fk__name')
        if self.request.GET.get('warehouse'):
            models = models.filter(warehouse_name_fk__in=self.request.GET.get('warehouse').split(','))
        if self.request.GET.get('date') and self.request.GET.get('date') != 'None':
            date = self.request.GET.get('date').split('to')
            date1 = date[0].split(' ')[0]
            date2 = date[1].split(' ')[1]
            date_start = datetime.strptime(date1, '%Y-%m-%d')
            date_end = datetime.strptime(date2, '%Y-%m-%d')
            models = models.filter(date__gte=date_start, date__lte=date_end)
        if self.request.GET.get('status'):
            if self.request.GET.get('status') == '0':
                pass
            else:
                models = models.filter(status=self.request.GET.get('status'))
        return models

    def get_context_data(self, **kwargs):
        context = super(WarehouseList, self).get_context_data(**kwargs)
        context['warehouse_get'] = self.request.GET.get('warehouse', '').strip(',')
        context['date_get'] = self.request.GET.get('date', '')
        context['status_get'] = self.request.GET.get('status', '')
        context['models_all'] = self.get_queryset().count()
        context['status_list'] = Warehouse.status_list
        warehouse = WarehouseName.objects.all().order_by('name')
        context['warehouse_name'] = warehouse
        warehouse_model_list = []
        for i in warehouse:
            w_list = {}
            models = self.get_queryset().filter(warehouse_name_fk=i)
            if models.count() > 0:
                w_list['name'] = i.name
                amount = 0
                for model in models:
                    amount = amount + model.amount
                w_list['amount'] = amount
                amount_use = 0
                for model in models:
                    amount_use = amount_use + model.amount_use
                w_list['amount_use'] = amount_use
                price = 0
                for model in models:
                    price = price + model.price
                w_list['price'] = price
                warehouse_model_list.append(w_list)
        context['warehouse_model_list'] = warehouse_model_list
        return context


class WarehouseCreate(BaseCreateView):
    template_name = 'administrator/warehouse/create.html'
    model = Warehouse
    fields = ('warehouse_name_fk', 'amount', 'price',)
    success_url = reverse_lazy('Administrator:warehouse_list')

    def get_context_data(self, **kwargs):
        context = super(WarehouseCreate, self).get_context_data(**kwargs)
        context['warehouse_name_models'] = WarehouseName.objects.all()
        return context


class WarehouseUpdate(BaseUpdateView):
    template_name = 'administrator/warehouse/update.html'
    model = Warehouse
    fields = ('warehouse_name_fk', 'amount', 'price',)
    success_url = reverse_lazy('Administrator:warehouse_list')

    def get_context_data(self, **kwargs):
        context = super(WarehouseUpdate, self).get_context_data(**kwargs)
        context['warehouse_name_models'] = WarehouseName.objects.all()
        return context


class WarehouseDelete(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        try:
            Warehouse.objects.get(pk=kwargs['pk']).delete()
            messages.success(request, "Maglumat pozuldy")
        except ObjectDoesNotExist:
            messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:warehouse_list'))


class WarehouseMultiDelete(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        for i in request.POST.getlist('item_id'):
            try:
                Warehouse.objects.get(pk=i).delete()
                messages.success(request, "Maglumat pozuldy")
            except ObjectDoesNotExist:
                messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:warehouse_list'))


class WarehouseExcelDownload(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Warehouse.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = 'Warehouse'
        sheet = wb.worksheets[0]
        # Ady
        cell = sheet.cell(row=1, column=1, value='Ady')
        cell.font = font_14_bold
        cell.border = border
        cell.alignment = center
        cell.fill = bg_color_light_green
        sheet.row_dimensions[1].height = 20
        sheet.column_dimensions[Alphabet(1)].width = 16
        wb.save(response)
        return response


class WarehouseExcelUpload(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        try:
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            shell = wb.active
            c = True
            i = 1
            while c:
                i += 1
                if shell.cell(row=i, column=1).value is None:
                    c = False
                    break
                else:
                    model = Warehouse()
                    model.name = shell.cell(row=i, column=1).value
                    model.save()
                messages.success(request, 'Maglumat girizildi')
            return HttpResponseRedirect(reverse('Administrator:warehouse_list'))
        except Exception as e:
            messages.error(request, f'Maglumat girizilmedi: {e}')
            return HttpResponseRedirect(reverse('Administrator:warehouse_list'))


class ProductList(BaseListView):
    template_name = 'administrator/product/list.html'

    def get_queryset(self):
        image_subquery = ProductImage.objects.filter(product_fk=OuterRef('pk')).values('image')[:1]
        return Product.objects.all().order_by('-id').annotate(image=Subquery(image_subquery),
                                                              display_percentage=Case(
                                                                  When(percentage=0, then=Value(False)),
                                                                  default=F('percentage'), output_field=IntegerField()))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_list'] = Product.status_list
        context['models_all'] = Product.objects.all().count()
        return context


class ProductCreate(BaseCreateView):
    template_name = 'administrator/product/create.html'
    model = Product
    fields = ('name',)
    success_url = reverse_lazy('Administrator:product_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_list'] = Product.status_list
        return context

    def post(self, request, *args, **kwargs):
        try:
            data = self.request.POST
            model = Product()
            model.name = data['name']
            model.cash_balance = data['cash_balance']
            model.description = data['description']
            model.price = data['price']
            model.is_active = data['status']
            model.description = data['description']
            if data['discount_option'] == '1':
                pass
            elif data['discount_option'] == '2':
                model.percentage = data['percentage']
            elif data['discount_option'] == '3':
                model.expensive_price = data['expensive_price']
            else:
                pass
            model.save()
            if self.request.FILES:
                for i in self.request.FILES.getlist('images'):
                    image_model = ProductImage()
                    image_model.product_fk = model
                    image_model.image = i
                    image_model.save()
            return JsonResponse({'message': 'success'}, status=200)
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)


class ProductUpdate(BaseUpdateView):
    template_name = 'administrator/product/update.html'
    model = Product
    fields = ('name',)
    success_url = reverse_lazy('Administrator:product_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_list'] = Product.status_list
        return context

    def post(self, request, *args, **kwargs):
        try:
            data = self.request.POST
            model = Product.objects.get(pk=self.kwargs['pk'])
            model.name = data['name']
            model.cash_balance = data['cash_balance']
            model.description = data['description']
            model.price = data['price']
            model.is_active = data['active']
            model.description = data['description']
            if data['discount_option'] == '1':
                pass
            elif data['discount_option'] == '2':
                model.percentage = data['percentage']
                model.expensive_price = 0
            elif data['discount_option'] == '3':
                model.percentage = 0
                model.expensive_price = data['expensive_price']
            else:
                pass
            model.save()
            if self.request.FILES:
                for i in self.request.FILES.getlist('images'):
                    image_model = ProductImage()
                    image_model.product_fk = model
                    image_model.image = i
                    image_model.save()
            return JsonResponse({'message': 'success'}, status=200)
        except Exception as e:
            return JsonResponse({'message': str(e)}, status=500)


class ProductDelete(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        try:
            Product.objects.get(pk=kwargs['pk']).delete()
            messages.success(request, "Maglumat pozuldy")
        except ObjectDoesNotExist:
            messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:product_list'))


class ProductMultiDelete(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        for i in request.POST.getlist('item_id'):
            try:
                Product.objects.get(pk=i).delete()
                messages.success(request, "Maglumat pozuldy")
            except ObjectDoesNotExist:
                messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:product_list'))


class ProductWarehouseList(BaseListView):
    template_name = 'administrator/product/warehouse/list.html'
    context_object_name = 'models'

    def get_queryset(self):
        active_warehouse = Warehouse.objects.filter(warehouse_name_fk=OuterRef('warehouse_name_fk'),
                                                    status='1').order_by('-date')
        one_price = ExpressionWrapper(F('price') / F('amount'),
                                      output_field=DecimalField(max_digits=10, decimal_places=2))
        subquery = active_warehouse.annotate(amount_div_price=one_price).values('amount_div_price')[:1]
        return ProductWarehouse.objects.filter(product_fk__id=self.kwargs['pk']).annotate(price=Subquery(subquery),
                                                                                          total=Value(
                                                                                              Subquery(subquery) * F(
                                                                                                  'amount')).value)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        product = Product.objects.get(pk=self.kwargs['pk'])
        context['product'] = product
        context['pk'] = self.kwargs['pk']
        context['count'] = self.get_queryset().count()
        context['my_cash_balance'] = (product.price - product.cost) * product.cash_balance / 100
        return context


class ProductCashBackView(BaseTemplateView):
    template_name = 'administrator/product/warehouse/cash_back.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['models'] = Product.objects.get(pk=self.kwargs['pk'])
        return context

    def post(self, request, *args, **kwargs):
        model = Product.objects.get(pk=self.kwargs['pk'])
        model.cash_balance = request.POST['cash_balance']
        model.save()
        return JsonResponse({'message': 'success'}, status=200)


class ProductWarehouseCreate(BaseTemplateView):
    template_name = 'administrator/product/warehouse/create.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['warehouse_models'] = WarehouseName.objects.all()
        context['pk'] = self.kwargs['pk']
        return context

    @staticmethod
    def post(request, *args, **kwargs):
        data = request.POST
        model = ProductWarehouse()
        model.product_fk_id = kwargs['pk']
        model.warehouse_name_fk_id = data['warehouse']
        model.amount = data['amount']
        model.save()
        total = 0
        active_warehouse = Warehouse.objects.filter(warehouse_name_fk=OuterRef('warehouse_name_fk'),
                                                    status='1').order_by('-date')
        one_price = ExpressionWrapper(F('price') / F('amount'),
                                      output_field=DecimalField(max_digits=10, decimal_places=2))
        subquery = active_warehouse.annotate(amount_div_price=one_price).values('amount_div_price')[:1]
        for i in ProductWarehouse.objects.filter(product_fk__id=kwargs['pk']).annotate(price=Subquery(subquery),
                                                                                       total=Value(
                                                                                               Subquery(subquery) * F(
                                                                                                       'amount')).value):
            try:
                total += i.total
            except:
                total += 0
        product = Product.objects.get(pk=kwargs['pk'])
        product.cost = total
        product.price_cost = product.price - total
        product.save()
        return HttpResponseRedirect(reverse('Administrator:product_warehouse_list', args=[model.product_fk.id]))


class ProductWarehouseUpdate(BaseTemplateView):
    template_name = 'administrator/product/warehouse/update.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['models'] = ProductWarehouse.objects.get(pk=self.kwargs['pk'])
        context['warehouse_models'] = WarehouseName.objects.all()
        return context

    @staticmethod
    def post(request, *args, **kwargs):
        data = request.POST
        model = ProductWarehouse.objects.get(pk=kwargs['pk'])
        model.warehouse_name_fk_id = data['warehouse']
        model.amount = data['amount']
        model.save()
        total = 0
        active_warehouse = Warehouse.objects.filter(warehouse_name_fk=OuterRef('warehouse_name_fk'),
                                                    status='1').order_by('-date')
        one_price = ExpressionWrapper(F('price') / F('amount'),
                                      output_field=DecimalField(max_digits=10, decimal_places=2))
        subquery = active_warehouse.annotate(amount_div_price=one_price).values('amount_div_price')[:1]
        for i in ProductWarehouse.objects.filter(product_fk=model.product_fk).annotate(price=Subquery(subquery),
                                                                                       total=Value(
                                                                                               Subquery(subquery) * F(
                                                                                                       'amount')).value):
            try:
                total += i.total
            except:
                total += 0
        product = Product.objects.get(pk=model.product_fk_id)
        product.cost = total
        product.price_cost = product.price - total
        product.save()
        return HttpResponseRedirect(reverse('Administrator:product_warehouse_list', args=[model.product_fk.id]))


class ProductWarehouseDelete(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        try:
            model = (ProductWarehouse.objects.get(pk=kwargs['pk']))
            pk = model.product_fk_id
            model.delete()
            total = 0
            active_warehouse = Warehouse.objects.filter(warehouse_name_fk=OuterRef('warehouse_name_fk'),
                                                        status='1').order_by('-date')
            one_price = ExpressionWrapper(F('price') / F('amount'),
                                          output_field=DecimalField(max_digits=10, decimal_places=2))
            subquery = active_warehouse.annotate(amount_div_price=one_price).values('amount_div_price')[:1]
            for i in ProductWarehouse.objects.filter(product_fk_id=pk).annotate(price=Subquery(subquery), total=Value(
                    Subquery(subquery) * F('amount')).value):
                try:
                    total += i.total
                except:
                    total += 0
            product = Product.objects.get(pk=pk)
            product.cost = total
            product.price_cost = product.price - total
            product.save()
            messages.success(request, "Maglumat pozuldy")
            return HttpResponseRedirect(reverse('Administrator:product_warehouse_list', args=[pk]))
        except ObjectDoesNotExist:
            messages.error(request, "Ýalňyşlyk ýüze çykdy")
            return HttpResponseRedirect(reverse('Administrator:product_list'))


class HomePageList(BaseListView):
    template_name = 'administrator/home_page/list.html'

    def get_queryset(self):
        models = HomePicture.objects.all()
        return models

    def get_context_data(self, **kwargs):
        context = super(HomePageList, self).get_context_data(**kwargs)
        context['models_all'] = self.get_queryset().count()
        return context


class HomePageCreate(BaseCreateView):
    template_name = 'administrator/home_page/create.html'
    model = HomePicture
    fields = '__all__'
    success_url = reverse_lazy('Administrator:homepage_list')
    success_message = '123'


class HomePageUpdate(BaseUpdateView):
    template_name = 'administrator/home_page/update.html'
    model = HomePicture
    fields = '__all__'
    success_url = reverse_lazy('Administrator:homepage_list')


class HomePageDelete(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        try:
            HomePicture.objects.get(pk=kwargs['pk']).delete()
            messages.success(request, "Maglumat pozuldy")
        except ObjectDoesNotExist:
            messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:homepage_list'))


class HomePageMultiDelete(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        for i in request.POST.getlist('item_id'):
            try:
                HomePicture.objects.get(pk=i).delete()
                messages.success(request, "Maglumat pozuldy")
            except ObjectDoesNotExist:
                messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:homepage_list'))


class HomePageExcelDownload(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="HomePage.xlsx"'
        wb = Workbook()
        ws = wb.active
        ws.title = 'HomePage'
        sheet = wb.worksheets[0]
        # Ady
        cell = sheet.cell(row=1, column=1, value='Ady')
        cell.font = font_14_bold
        cell.border = border
        cell.alignment = center
        cell.fill = bg_color_light_green
        sheet.row_dimensions[1].height = 20
        sheet.column_dimensions[Alphabet(1)].width = 16
        wb.save(response)
        return response


class HomePageExcelUpload(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        try:
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            shell = wb.active
            c = True
            i = 1
            while c:
                i += 1
                if shell.cell(row=i, column=1).value is None:
                    c = False
                    break
                else:
                    model = HomePicture()
                    model.name = shell.cell(row=i, column=1).value
                    model.save()
                messages.success(request, 'Maglumat girizildi')
            return HttpResponseRedirect(reverse('Administrator:warehouse_name_list'))
        except Exception as e:
            messages.error(request, f'Maglumat girizilmedi: {e}')
            return HttpResponseRedirect(reverse('Administrator:warehouse_name_list'))


class UsersList(BaseListView):
    template_name = 'administrator/users/list.html'

    def get_queryset(self):
        models = Users.objects.all()
        if self.request.GET.get('user_type') and self.request.GET.get('user_type') != '0':
            models = models.filter(user_type=self.request.GET.get('user_type'))
        if self.request.GET.get('search'):
            models = models.filter(phone_number__icontains=self.request.GET.get('search'))
        return models

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_type_value'] = self.request.GET.get('user_type', 0)
        context['user_type_list'] = Users.user_type_list
        context['models_all'] = self.get_queryset().count()
        return context


class UsersCreate(BaseCreateView):
    template_name = 'administrator/users/create.html'
    model = Users
    fields = ('phone_number',)

    def post(self, request, *args, **kwargs):
        user = Users()
        try:
            user.username = request.POST.get('username')
            user.phone_number = request.POST.get('phone_number')
            user.set_password(request.POST.get('password'))
            user.user_type = request.POST.get('user_type')
            user.save()
        except Exception as e:
            print(e)
        return HttpResponseRedirect(reverse('Administrator:users_list'))


class UsersUpdate(BaseUpdateView):
    template_name = 'administrator/users/update.html'
    model = Users
    fields = ('phone_number',)
    success_url = reverse_lazy('Administrator:users_list')

    def post(self, request, *args, **kwargs):
        try:
            user = Users.objects.get(id=self.kwargs.get('pk'))
            user.username = request.POST.get('username')
            user.phone_number = request.POST.get('phone_number')
            user.user_type = request.POST.get('user_type')
            user.save()
        except Exception as e:
            print(e)
        return HttpResponseRedirect(reverse('Administrator:users_list'))


class UsersDelete(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        try:
            Users.objects.get(pk=kwargs['pk']).delete()
            messages.success(request, "Maglumat pozuldy")
        except ObjectDoesNotExist:
            messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:users_list'))


class UsersMultiDelete(BaseView):
    @staticmethod
    def post(request, *args, **kwargs):
        for i in request.POST.getlist('item_id'):
            try:
                Users.objects.get(pk=i).delete()
                messages.success(request, "Maglumat pozuldy")
            except ObjectDoesNotExist:
                messages.error(request, "Ýalňyşlyk ýüze çykdy")
        return HttpResponseRedirect(reverse('Administrator:users_list'))


class OrderListView(BaseListView):
    template_name = 'administrator/order/list.html'
    context_object_name = 'models'

    def get_queryset(self):
        models = Order.objects.all().order_by('-date').order_by('order_state')
        if self.request.GET.get('search'):
            models = models.filter(name__icontains=self.request.GET.get('search'))
        if self.request.GET.get('status'):
            models = models.filter(order_state__in=self.request.GET.get('status').split(','))
        return models

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['order_status_list'] = Order.order_status_list
        context['models_all'] = self.get_queryset().count()
        context['order_status_list'] = Order.order_status_list
        context['payment_type_list'] = Order.payment_type_list
        context['order_status'] = self.request.GET.get('status', '').split(',')
        return context


class OrderDetailView(BaseTemplateView):
    template_name = 'administrator/order/detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        models = Order.objects.get(name=self.kwargs['code'])
        context['models'] = models
        context['order_status_list'] = Order.order_status_list
        context['payment_type_list'] = Order.payment_type_list
        image_subquery = ProductImage.objects.filter(product_fk=OuterRef('product_fk')).values('image')[:1]
        context['order_models'] = OrderItem.objects.filter(order=models).annotate(image=Subquery(image_subquery),
                                                                                  total=Value(
                                                                                      F('price') * F('quantity')).value)
        return context


class OrderRegister(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        order = Order.objects.get(id=kwargs['pk'])
        order.order_state = 2
        order.save()
        return HttpResponseRedirect(reverse('Administrator:order_detail_list', args={order.name}))


class OrderCancel(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):
        print(kwargs['pk'])
        order = Order.objects.get(id=kwargs['pk'])
        order.order_state = 5
        order.save()
        return HttpResponseRedirect(reverse('Administrator:order_detail_list', args={order.name}))


def calc_cash_balance(product):
    money = 0.0
    money = (product.price - product.cost) * product.cash_balance / 100
    return money


class OrderSuccess(BaseView):
    @staticmethod
    def get(request, *args, **kwargs):

        order = Order.objects.get(id=kwargs['pk'])
        order.order_state = 3
        order.save()
        if order.user:
            for i in OrderItem.objects.filter(order=order):
                for c in range(0, i.quantity):
                    money = calc_cash_balance(i.product_fk)
                    payment = Payment()
                    payment.user_fk = order.user
                    payment.order = i
                    payment.money = money
                    payment.save()
                    users = Users.objects.get(pk=order.user_id)
                    users.wallet += money
                    users.save()
                    for w in ProductWarehouse.objects.filter(product_fk=i.product_fk):
                        print(f'{w.product_fk.name} + {w.warehouse_name_fk.name} == {w.amount}')
                        active_warehouse = Warehouse.objects.filter(warehouse_name_fk=w.warehouse_name_fk,
                                                                    status='1').order_by('-date')[:1].first()
                        print(f'{active_warehouse.warehouse_name_fk.name} + {active_warehouse.amount_use}')
                        if active_warehouse.amount_use + w.amount > active_warehouse.amount:
                            print('-----------------------------------')
                            amount_use = active_warehouse.amount - active_warehouse.amount_use
                            active_warehouse.active_warehouse.amount_use += amount_use
                            amount = active_warehouse.amount_use + w.amount - active_warehouse.amount
                            active_warehouse.status = '2'
                            active_warehouse.save()
                            active_warehouse = Warehouse.objects.filter(warehouse_name_fk=w.warehouse_name_fk,
                                                                        status='1').order_by('-date')[:1].first()
                            if active_warehouse:
                                active_warehouse.amount_use += amount
                                active_warehouse.save()
                            else:
                                # Name etmeli eger ammarda produkta gutarsa? TEKLIP!
                                pass
                        elif active_warehouse.amount_use + w.amount == active_warehouse.amount:
                            print('-----------------------------------')
                            active_warehouse.amount_use += w.amount
                            active_warehouse.status = '2'
                            active_warehouse.save()
                        else:
                            print('-----------------------------------')
                            active_warehouse.amount_use += w.amount
                            active_warehouse.save()
        return HttpResponseRedirect(reverse('Administrator:order_detail_list', args={order.name}))
